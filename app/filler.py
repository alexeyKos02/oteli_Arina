"""
Заполнение фиксированного Excel-шаблона извлечёнными ставками + отчёт.

Пишем ТОЛЬКО во входные ячейки "Bed prices". Формулы комбинаций в шаблоне не трогаем —
openpyxl сохраняет их (в т.ч. массивные) при повторной записи, если ячейки не изменялись.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from dataclasses import dataclass, field

import openpyxl

from .mapping import (
    PERIOD_COLUMNS,
    N_PERIODS,
    OFFSET_ADULT_MB,
    OFFSET_ADULT_EXB,
    OFFSET_CHILD312_MB,
    OFFSET_CHILD312_EXB,
    OFFSET_SINGLE_1A,
    OFFSET_PERIOD_START,
)
from .parser import Category, CODE_RE

# Множитель одиночного размещения (1A) в формуле шаблона: Single = Adult MB * 1.7.
# Используем для кросс-проверки: parsed "Single net" должен совпасть -> высокая достоверность.
SINGLE_MULTIPLIER = 1.7


@dataclass
class RoomReport:
    code: str
    name: str
    matched: bool
    written: dict[str, list[float]] = field(default_factory=dict)  # field -> 6 значений
    note: str = ""
    confidence: float = 0.0          # 0..1
    level: str = "—"                 # high | medium | low | —
    flags: list[str] = field(default_factory=list)  # причины снижения достоверности


@dataclass
class FillReport:
    rooms: list[RoomReport] = field(default_factory=list)
    unmatched_pdf_codes: list[str] = field(default_factory=list)  # есть в PDF, нет в шаблоне
    empty_template_codes: list[str] = field(default_factory=list)  # есть в шаблоне, нет в PDF
    warnings: list[str] = field(default_factory=list)

    @property
    def matched_count(self) -> int:
        return sum(1 for r in self.rooms if r.matched)


def _template_room_blocks(ws) -> list[tuple[int, str, str]]:
    """Список (row, full_name, code) для всех блоков комнат шаблона."""
    blocks = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Room":
            name = str(ws.cell(r, 2).value or "")
            m = CODE_RE.search(name)
            code = m.group(1) if m else ""
            blocks.append((r, name, code))
    return blocks


def _write_row(ws, row: int, values: list[float], cols: list[int]) -> None:
    for col, val in zip(cols, values):
        ws.cell(row=row, column=col, value=val)


def _block_end(ws, row: int, room_rows: list[int]) -> int:
    nxt = [r for r in room_rows if r > row]
    return min(nxt) if nxt else ws.max_row + 1


def _has_child_combo(ws, row: int, room_rows: list[int]) -> bool:
    """Есть ли в блоке комбинации с ребёнком (метка содержит 'c12') — признак семейной комнаты."""
    for r in range(row, _block_end(ws, row, room_rows)):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and "c12" in a.lower():
            return True
    return False


def fill_labeled_blocks(ws, row: int, cat: Category, cols: list[int], room_rows) -> dict:
    """Запись для формата labeled_blocks (Adult MB/ExB, Child 3-12, Single-надбавка)."""
    written: dict[str, list[float]] = {}
    adult_mb = cat.rates.get("adult_mb")
    adult_exb = cat.rates.get("adult_exb")
    child_exb = cat.rates.get("child312_exb")
    single = cat.rates.get("single")

    if adult_mb:
        _write_row(ws, row + OFFSET_ADULT_MB, adult_mb, cols)
        written["Adult MB"] = adult_mb
    # 1A (Single): формула шаблона не учитывает надбавку за одноместное размещение —
    # пишем "Single net" из PDF поверх формулы.
    if single:
        _write_row(ws, row + OFFSET_SINGLE_1A, single, cols)
        written["1A (Single)"] = single
    if adult_exb:
        _write_row(ws, row + OFFSET_ADULT_EXB, adult_exb, cols)
        written["Adult ExB"] = adult_exb
    # Child 3-12 заполняем только если PDF даёт эту строку.
    if child_exb:
        _write_row(ws, row + OFFSET_CHILD312_EXB, child_exb, cols)
        written["Child 3-12 ExB"] = child_exb
        if adult_mb:  # Child 3-12 MB = Adult MB (правило подтверждено на эталоне)
            _write_row(ws, row + OFFSET_CHILD312_MB, adult_mb, cols)
            written["Child 3-12 MB"] = adult_mb
    return written


def fill_rate_matrix(ws, row: int, cat: Category, cols: list[int], room_rows) -> dict:
    """
    Запись для формата rate_matrix: только базовый Adult MB. Всё прочее (1A=×1.6,
    ExB=×0.85, Child ExB=×0.5, 2A) — формулы шаблона. Для семейных комнат Child MB
    тоже = Adult MB (эти комбинации на него ссылаются).
    """
    written: dict[str, list[float]] = {}
    adult_mb = cat.rates.get("adult_mb")
    if adult_mb:
        _write_row(ws, row + OFFSET_ADULT_MB, adult_mb, cols)
        written["Adult MB"] = adult_mb
        if _has_child_combo(ws, row, room_rows):
            _write_row(ws, row + OFFSET_CHILD312_MB, adult_mb, cols)
            written["Child 2-12 MB"] = adult_mb
    return written


def fill_template(template_path: str, categories: list[Category], profile) -> tuple[bytes, FillReport]:
    """
    Заполнить шаблон согласно профилю формата. Возвращает (xlsx_bytes, report).
    profile: объект с полями period_columns, n_periods, required_keys,
             single_multiplier, fill_room(ws,row,cat,cols,room_rows)->dict.
    """
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb.active
    cols = profile.period_columns

    blocks = _template_room_blocks(ws)
    code_to_block = {code: (row, name) for row, name, code in blocks if code}
    room_rows = [row for row, _, _ in blocks]

    # code -> Category (в labeled_blocks King и Double делят одну категорию)
    code_to_cat: dict[str, Category] = {}
    all_pdf_codes: list[str] = []
    for cat in categories:
        for code in cat.codes:
            all_pdf_codes.append(code)
            code_to_cat[code] = cat

    report = FillReport()

    for code, (row, name) in sorted(code_to_block.items(), key=lambda x: x[1][0]):
        cat = code_to_cat.get(code)
        if cat is None:
            report.rooms.append(RoomReport(code=code, name=name, matched=False,
                                           note="нет в PDF (не сопоставлено)"))
            report.empty_template_codes.append(code)
            continue

        rr = RoomReport(code=code, name=name, matched=True)
        rr.written = profile.fill_room(ws, row, cat, cols, room_rows)
        if not rr.written:
            rr.note = "категория найдена, но ставки не извлеклись"

        _score_room(ws, row, cat, rr, profile)
        if rr.level in ("low", "medium"):
            report.warnings.append(f"{code}: достоверность {rr.level} — {'; '.join(rr.flags)}")
        report.rooms.append(rr)

    # Коды из PDF, которых нет в шаблоне.
    seen = set()
    for code in all_pdf_codes:
        if code not in code_to_block and code not in seen:
            seen.add(code)
            report.unmatched_pdf_codes.append(code)

    # openpyxl не сохраняет кэш значений формул -> просим Excel пересчитать при открытии,
    # иначе комбинации покажут 0 до ручного пересчёта.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), report


def _parse_any_date(value) -> date | None:
    """Привести дату из PDF ('3-Jan-27') или шаблона ('03/01/2027', datetime) к date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d-%b-%y", "%d/%m/%Y", "%d/%m/%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# Метки обязательных ключей для сообщений
_KEY_LABELS = {"adult_mb": "Adult MB", "adult_exb": "Adult ExB", "child312_exb": "Child ExB"}


def _score_room(ws, row: int, cat: Category, rr: RoomReport, profile) -> None:
    """
    Оценить достоверность извлечения комнаты. Заполняет rr.confidence/level/flags.
    Параметры берутся из профиля: число периодов, обязательные ключи, множитель
    кросс-проверки Single (если формат его содержит).
    """
    n_periods = profile.n_periods
    score = 1.0
    flags: list[str] = []

    # 1) число периодов
    n = len(cat.period_starts)
    if n != n_periods:
        score -= 0.4
        flags.append(f"{n} периодов вместо {n_periods}")

    # 2) полнота обязательных строк
    for key in profile.required_keys:
        label = _KEY_LABELS.get(key, key)
        vals = cat.rates.get(key)
        if not vals:
            score -= 0.3
            flags.append(f"нет строки «{label}»")
        elif len(vals) != n_periods:
            score -= 0.2
            flags.append(f"«{label}»: {len(vals)} значений вместо {n_periods}")

    # 3) сверка дат периодов с шаблоном (пропускается, если даты не парсятся, напр. диапазоны)
    tmpl_dates = [_parse_any_date(ws.cell(row + OFFSET_PERIOD_START, c).value)
                  for c in profile.period_columns]
    pdf_dates = [_parse_any_date(d) for d in cat.period_starts[:n_periods]]
    mism = sum(1 for a, b in zip(tmpl_dates, pdf_dates)
               if a and b and a != b)
    if mism:
        score -= min(0.3, 0.1 * mism)
        flags.append(f"даты периодов расходятся с шаблоном ({mism})")

    # 4) кросс-проверка Single ≈ Adult MB * множитель (только для форматов с Single)
    mult = getattr(profile, "single_multiplier", None)
    single = cat.rates.get("single")
    adult_mb = cat.rates.get("adult_mb")
    if mult and single and adult_mb and len(single) == len(adult_mb):
        bad = sum(1 for s, a in zip(single, adult_mb)
                  if abs(s - a * mult) > 1.0)
        if bad:
            score -= min(0.3, 0.1 * bad)
            flags.append(f"Single ≠ AdultMB×{mult} в {bad} периодах")

    # 5) положительность значений
    for key, vals in cat.rates.items():
        if any((v is None or v <= 0) for v in vals):
            score -= 0.1
            flags.append(f"нечисловые/неположительные значения в «{key}»")
            break

    score = max(0.0, min(1.0, score))
    rr.confidence = round(score, 2)
    rr.flags = flags
    rr.level = "high" if score >= 0.9 else ("medium" if score >= 0.6 else "low")
